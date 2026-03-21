# xlsx ingestion (background jobs)

import os
import openpyxl
from celery import shared_task
from django.conf import settings
from .models import Customer, Loan
from datetime import date


@shared_task
def ingest_customer_data():
    path = os.path.join(settings.BASE_DIR, "customer_data.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Read the data from the worksheet and create customers
    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            customer_id,
            first_name,
            last_name,
            phone_number,
            monthly_salary,
            age,
            approved_limit,
            current_debt,
        ) = row

        Customer.objects.get_or_create(
            customer_id=customer_id,
            defaults={
                "first_name": first_name,
                "last_name": last_name,
                "phone_number": str(phone_number),
                "monthly_salary": monthly_salary,
                "age": age,
                "approved_limit": approved_limit,
                "current_debt": current_debt or 0,
            },
        )


@shared_task
def ingest_loan_data():
    path = os.path.join(settings.BASE_DIR, "loan_data.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            customer_id,
            loan_id,
            loan_amount,
            tenure,
            interest_rate,
            monthly_repayment,
            emis_paid_on_time,
            start_date,
            end_date,
        ) = row

        try:
            customer = Customer.objects.get(customer_id=customer_id)
        except Customer.DoesNotExist:
            continue

        Loan.objects.get_or_create(
            loan_id=loan_id,
            defaults={
                "customer": customer,
                "loan_amount": loan_amount,
                "tenure": tenure,
                "interest_rate": interest_rate,
                "monthly_repayment": monthly_repayment,
                "emis_paid_on_time": emis_paid_on_time,
                "start_date": start_date,
                "end_date": end_date,
            },
        )
